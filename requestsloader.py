# -*- coding: utf-8 -*-

import logging, datetime

from openpyxl import load_workbook

from vars_setting import PARMFILE_NAME, DEF_SHEET_NAME

#singleton mode via decorator 
def singleton(cls): 
	_instance = {}
	def _warper(*args, **kargs):
		if cls not in _instance:
			_instance[cls] = cls(*args, **kargs)
		return _instance[cls]
	return _warper

# Class definition

class RequestsLoader(object):
	"""docstring for ParameterLoader
	load the parameter file and retrun the list of request
	"""
	def __init__(self):
		self.parameter_filename = ''
		self.sheetname = ''
		self.wb = None
		self.sheet = None

	def load_workbook(self, parameter_filename=PARMFILE_NAME, sheetname=DEF_SHEET_NAME):

		self.parameter_filename = parameter_filename
		try:
			self.wb = load_workbook(self.parameter_filename)
		except Exception as e:
			raise

		try:
			self.sheet = self.wb[sheetname]
		except Exception as e:
			raise

	def get_records(self):

		if self.sheet is None:
			logging.error('Please init the loadrequest object')
			return None

		records = []
		for row in range(1, self.sheet.max_row+1):
			records.append(self.sheet[row])
		return records

	def get_requests(self):

		if self.sheet is None:
			logging.error('Please init the loadrequest object')
			return None

		records_list = []
		title_value = [col.value for col in self.sheet[1]]

		logging.debug('The maxrow of parameter file is %s' %self.sheet.max_row )
		
		for row in range(2, self.sheet.max_row+1):
			row_value = [ col.value for col in self.sheet[row]]
			records_list.append(dict(zip(title_value, row_value)))
		return records_list

	def get_requests_str(self):

		if self.sheet is None:
			logging.error('Please init the loadrequest object')
			return None

		records_list = []
		title_value = [col.value for col in self.sheet[1]]

		logging.debug('The maxrow of parameter file is %s' %self.sheet.max_row )
		
		for row in range(2, self.sheet.max_row+1):
			row_value = []
			for col in self.sheet[row]:
				if isinstance(col.value, datetime.datetime):
					col.value = '-'.join([str(col.value.year), str(col.value.month), str(col.value.day)])
				row_value.append(col.value)
				
			records_list.append(dict(zip(title_value, row_value)))
		return records_list


if __name__ == '__main__':
	loadrequest = RequestsLoader()
	loadrequest.load_workbook()
	requests = loadrequest.get_requests_str()

	for request in requests:
		for key, value in request.items():
			print('%s : %s, type is %s' %(key, value, type(value)))